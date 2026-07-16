import uuid
from datetime import date, datetime, timedelta
from decimal import Decimal
from io import BytesIO
from unittest.mock import Mock, patch

from django.contrib.auth import get_user_model
from django.test import TestCase, override_settings
from django.utils import timezone
from openpyxl import load_workbook

from .korona import KoronaError
from .models import (
    BulkOrderList,
    DeferredReceipt,
    Delivery,
    DeliveryAsset,
    DeliveryEvent,
    DeliveryKeyword,
    OrderItemTransfer,
    OrderList,
    OrderListItem,
    Product,
    ProductCode,
    ProductStock,
    ApiRequestLog,
    ReceiptSaleLine,
    SalesDailySummary,
    ServiceControl,
    Store,
    SystemSetting,
    SystemLog,
    SyncState,
    SyncRun,
)
from .services import (
    STOCK_SYNC_ENTITY,
    _apply_receipt,
    rebuild_all_monthly_needs,
    reconcile_monthly_needs,
    recalculate_monthly_need,
    recalculate_monthly_needs_for_pairs,
    recalculate_stale_monthly_needs,
    replay_deferred_receipts,
    reconcile_store_stocks,
    sync_monthly_receipt_history,
    sync_receipts,
    sync_store_stocks_incremental,
)
from .serializers import supplier_short_name
from .tasks import (
    INTERRUPTED_MESSAGE,
    MONTHLY_CONFLICT_REASON,
    reconcile_stocks_task,
    reconcile_monthly_totals_task,
    recover_interrupted_runs,
    resolve_worker_interruptions,
    sync_receipts_task,
)


class FakeStoreStockClient:
    def __init__(self, pages):
        self.pages = pages
        self.calls = []

    def organizational_unit_product_stocks(self, store_id, revision=None, page_size=None):
        self.calls.append({"store_id": store_id, "revision": revision, "page_size": page_size})
        return iter(self.pages)


class RetiredWarehouseClient:
    def organizational_unit_product_stocks(self, store_id, revision=None, page_size=None):
        raise KoronaError(
            "organizational unit contains no warehouse",
            status_code=404,
            error_code="CONDITION_MISMATCH",
        )

    def organizational_unit(self, store_id):
        return {
            "id": str(store_id),
            "number": "3210",
            "name": "Main",
            "active": True,
            "warehouse": False,
            "revision": 31,
        }


class StoreStockSyncTests(TestCase):
    def setUp(self):
        self.store = Store.objects.create(
            korona_id=uuid.uuid4(), number="3210", name="Main", is_warehouse=True
        )
        self.product = Product.objects.create(
            korona_id=uuid.uuid4(), number="100", name="Test Bottle", normalized_name="testbottle"
        )

    def stock_row(self, product=None, *, actual=7, revision=20):
        return {
            "product": {"id": str((product or self.product).korona_id)},
            "warehouse": {"id": str(self.store.korona_id)},
            "revision": revision,
            "amount": {
                "actual": actual,
                "ordered": 2,
                "lent": 1,
                "reorderLevel": 3,
                "maxLevel": 12,
            },
        }

    def test_incremental_sync_uses_store_cursor_and_bulk_updates_stock(self):
        ProductStock.objects.create(product=self.product, store=self.store, actual=99, revision=10)
        SyncState.objects.create(entity=STOCK_SYNC_ENTITY, store=self.store, last_revision=10)
        client = FakeStoreStockClient([([self.stock_row()], 20)])

        counts = sync_store_stocks_incremental(self.store, client)

        stock = ProductStock.objects.get(product=self.product, store=self.store)
        state = SyncState.objects.get(entity=STOCK_SYNC_ENTITY, store=self.store)
        self.assertEqual(stock.actual, Decimal("7"))
        self.assertEqual(stock.ordered, Decimal("2"))
        self.assertEqual(stock.lent, Decimal("1"))
        self.assertEqual(stock.revision, 20)
        self.assertEqual(state.last_revision, 20)
        self.assertEqual(client.calls[0]["revision"], 10)
        self.assertEqual(
            counts,
            {"seen": 1, "created": 0, "updated": 1, "unchanged": 0, "deferred": 0},
        )

    def test_incremental_sync_does_not_report_an_unchanged_row_as_updated(self):
        ProductStock.objects.create(
            product=self.product,
            store=self.store,
            actual=7,
            ordered=2,
            lent=1,
            reorder_level=3,
            max_level=12,
            revision=10,
        )
        SyncState.objects.create(entity=STOCK_SYNC_ENTITY, store=self.store, last_revision=10)
        client = FakeStoreStockClient([([self.stock_row(revision=20)], 20)])

        counts = sync_store_stocks_incremental(self.store, client)

        self.assertEqual(counts["created"], 0)
        self.assertEqual(counts["updated"], 0)
        self.assertEqual(counts["unchanged"], 1)
        self.assertEqual(ProductStock.objects.get(product=self.product, store=self.store).revision, 20)

    def test_incremental_cursor_does_not_advance_past_unknown_product(self):
        SyncState.objects.create(entity=STOCK_SYNC_ENTITY, store=self.store, last_revision=10)
        unknown = Product(
            korona_id=uuid.uuid4(), number="missing", name="Missing", normalized_name="missing"
        )
        client = FakeStoreStockClient([([self.stock_row(unknown, revision=21)], 21)])

        counts = sync_store_stocks_incremental(self.store, client)

        state = SyncState.objects.get(entity=STOCK_SYNC_ENTITY, store=self.store)
        self.assertEqual(state.last_revision, 10)
        self.assertEqual(counts["deferred"], 1)

    def test_full_reconciliation_resets_stock_not_returned_by_korona(self):
        missing_product = Product.objects.create(
            korona_id=uuid.uuid4(), number="200", name="Old Stock", normalized_name="oldstock"
        )
        ProductStock.objects.create(
            product=missing_product,
            store=self.store,
            actual=8,
            ordered=4,
            lent=2,
            reorder_level=1,
            max_level=10,
        )
        client = FakeStoreStockClient([([self.stock_row(actual=5, revision=30)], 30)])

        counts = reconcile_store_stocks(self.store, client)

        stale = ProductStock.objects.get(product=missing_product, store=self.store)
        current = ProductStock.objects.get(product=self.product, store=self.store)
        state = SyncState.objects.get(entity=STOCK_SYNC_ENTITY, store=self.store)
        self.assertEqual(current.actual, Decimal("5"))
        self.assertEqual(stale.actual, Decimal("0"))
        self.assertEqual(stale.ordered, Decimal("0"))
        self.assertEqual(stale.lent, Decimal("0"))
        self.assertEqual(state.last_revision, 30)
        self.assertEqual(counts["created"], 1)
        self.assertEqual(counts["updated"], 1)

    def test_incremental_sync_retires_store_that_is_no_longer_a_warehouse(self):
        ProductStock.objects.create(product=self.product, store=self.store, actual=8, revision=30)
        SyncState.objects.create(entity=STOCK_SYNC_ENTITY, store=self.store, last_revision=30)

        counts = sync_store_stocks_incremental(self.store, RetiredWarehouseClient())

        self.store.refresh_from_db()
        self.assertTrue(self.store.active)
        self.assertFalse(self.store.is_warehouse)
        self.assertEqual(self.store.revision, 31)
        self.assertFalse(ProductStock.objects.filter(store=self.store).exists())
        self.assertFalse(SyncState.objects.filter(entity=STOCK_SYNC_ENTITY, store=self.store).exists())
        self.assertEqual(counts, {"seen": 0, "created": 0, "updated": 0, "deferred": 0, "retired": 1})

    @patch("orders.tasks.reconcile_stocks", return_value={"seen": 0, "created": 0, "updated": 0})
    def test_nightly_task_ignores_interval_gate_but_keeps_fixed_schedule(self, reconcile):
        control, _ = ServiceControl.objects.get_or_create(service_name="stock_reconciliation")
        control.enabled = True
        control.next_run_at = timezone.now() + timedelta(hours=12)
        control.save()

        with patch.dict(
            "orders.tasks.SERVICES",
            {"stock_reconciliation": (reconcile, 86400)},
        ):
            result = reconcile_stocks_task()

        control.refresh_from_db()
        reconcile.assert_called_once_with()
        self.assertEqual(result["seen"], 0)
        self.assertIsNone(control.next_run_at)

    def test_expired_running_job_is_reported_as_interrupted(self):
        started = timezone.now() - timedelta(minutes=20)
        control = ServiceControl.objects.create(
            service_name="monthly_reconciliation",
            status=ServiceControl.Status.RUNNING,
            locked_until=timezone.now() - timedelta(minutes=5),
        )
        run = SyncRun.objects.create(
            job_name="monthly_reconciliation",
            status=SyncRun.Status.RUNNING,
            started_at=started,
        )

        self.assertEqual(recover_interrupted_runs(), 1)

        control.refresh_from_db()
        run.refresh_from_db()
        self.assertEqual(control.status, ServiceControl.Status.ERROR)
        self.assertIsNone(control.locked_until)
        self.assertEqual(control.last_error, INTERRUPTED_MESSAGE)
        self.assertEqual(run.status, SyncRun.Status.ERROR)
        self.assertEqual(run.error_message, INTERRUPTED_MESSAGE)
        self.assertGreaterEqual(run.duration_ms, 20 * 60 * 1000)

    def test_successful_retry_marks_matching_worker_interruptions_resolved(self):
        matching = SystemLog.objects.create(
            level="ERROR",
            source="sync.worker",
            message=INTERRUPTED_MESSAGE,
            context={"services": ["monthly_reconciliation"]},
        )
        unrelated = SystemLog.objects.create(
            level="ERROR",
            source="sync.worker",
            message=INTERRUPTED_MESSAGE,
            context={"services": ["stock_reconciliation"]},
        )
        resolved_at = timezone.now()

        self.assertEqual(resolve_worker_interruptions("monthly_reconciliation", resolved_at), 1)

        matching.refresh_from_db()
        unrelated.refresh_from_db()
        self.assertEqual(matching.context["resolved_by"], "monthly_reconciliation")
        self.assertEqual(matching.context["resolved_at"], resolved_at.isoformat())
        self.assertNotIn("resolved_at", unrelated.context)

    def test_monthly_reconciliation_waits_then_receipt_task_dispatches_it(self):
        receipt_control = ServiceControl.objects.create(
            service_name="receipts",
            status=ServiceControl.Status.RUNNING,
            locked_until=timezone.now() + timedelta(minutes=5),
        )
        monthly_control = ServiceControl.objects.create(
            service_name="monthly_reconciliation",
            status=ServiceControl.Status.QUEUED,
        )
        monthly_work = Mock(return_value={"seen": 1, "created": 1, "updated": 0})

        with patch.dict(
            "orders.tasks.SERVICES",
            {"monthly_reconciliation": (monthly_work, 86400)},
        ):
            waiting = reconcile_monthly_totals_task(force=True)

        monthly_control.refresh_from_db()
        monthly_work.assert_not_called()
        self.assertEqual(waiting["reason"], MONTHLY_CONFLICT_REASON)
        self.assertTrue(waiting["queued"])
        self.assertEqual(monthly_control.status, ServiceControl.Status.QUEUED)

        receipt_control.status = ServiceControl.Status.IDLE
        receipt_control.locked_until = None
        receipt_control.save(update_fields=["status", "locked_until", "updated_at"])
        receipt_work = Mock(return_value={"seen": 1, "created": 0, "updated": 0})
        with patch("orders.tasks.reconcile_monthly_totals_task.delay") as delay, patch.dict(
            "orders.tasks.SERVICES", {"receipts": (receipt_work, 120)}
        ):
            receipt_result = sync_receipts_task(force=True)

        receipt_work.assert_not_called()
        self.assertEqual(receipt_result["reason"], MONTHLY_CONFLICT_REASON)
        delay.assert_called_once_with(force=True)

        with patch.dict(
            "orders.tasks.SERVICES",
            {"monthly_reconciliation": (monthly_work, 86400)},
        ):
            completed = reconcile_monthly_totals_task(force=True)

        monthly_control.refresh_from_db()
        monthly_work.assert_called_once_with()
        self.assertEqual(completed["seen"], 1)
        self.assertEqual(monthly_control.status, ServiceControl.Status.IDLE)

    def test_skipped_monthly_job_clears_a_stale_queued_status(self):
        control = ServiceControl.objects.create(
            service_name="monthly_reconciliation",
            status=ServiceControl.Status.QUEUED,
            locked_until=timezone.now() + timedelta(minutes=5),
        )

        result = reconcile_monthly_totals_task(force=True)

        control.refresh_from_db()
        self.assertEqual(result["reason"], "already running")
        self.assertEqual(control.status, ServiceControl.Status.IDLE)
        self.assertIsNone(control.locked_until)


class ReceiptSyncTests(TestCase):
    def setUp(self):
        self.store = Store.objects.create(korona_id=uuid.uuid4(), number="1", name="Main")
        self.product = Product.objects.create(
            korona_id=uuid.uuid4(), number="100", name="Test Bottle", normalized_name="testbottle"
        )
        self.receipt_id = uuid.uuid4()

    def receipt(self, quantity, revision=1, cancelled=False):
        return {
            "id": str(self.receipt_id),
            "revision": revision,
            "bookingTime": "2026-07-10T12:00:00-04:00",
            "organizationalUnit": {"id": str(self.store.korona_id)},
            "cancelled": cancelled,
            "items": [{"product": {"id": str(self.product.korona_id)}, "quantity": quantity}],
        }

    def test_receipt_revision_applies_only_delta(self):
        affected, counts = set(), {"created": 0, "updated": 0}
        _apply_receipt(self.receipt(2), affected, counts)
        _apply_receipt(self.receipt(5, revision=2), affected, counts)
        summary = SalesDailySummary.objects.get(store=self.store, product=self.product)
        self.assertEqual(summary.quantity_sold, Decimal("5"))
        self.assertEqual(summary.receipts_count, 1)
        self.assertEqual(ReceiptSaleLine.objects.get().quantity, Decimal("5"))

    def test_replaying_same_revision_is_idempotent(self):
        affected, counts = set(), {"created": 0, "updated": 0}
        receipt = self.receipt(4, revision=7)

        _apply_receipt(receipt, affected, counts)
        _apply_receipt(receipt, affected, counts)

        summary = SalesDailySummary.objects.get(store=self.store, product=self.product)
        self.assertEqual(summary.quantity_sold, Decimal("4"))
        self.assertEqual(summary.receipts_count, 1)
        self.assertEqual(ReceiptSaleLine.objects.count(), 1)

    def test_unknown_product_receipt_is_replayed_after_catalog_sync(self):
        unknown_product_id = uuid.uuid4()
        receipt = self.receipt(6, revision=8)
        receipt["items"][0]["product"]["id"] = str(unknown_product_id)
        affected, counts = set(), {"created": 0, "updated": 0}

        _apply_receipt(receipt, affected, counts)

        self.assertEqual(DeferredReceipt.objects.count(), 1)
        self.assertFalse(ReceiptSaleLine.objects.filter(receipt_id=self.receipt_id).exists())

        product = Product.objects.create(
            korona_id=unknown_product_id,
            number="NEW-1",
            name="New Product",
            normalized_name="newproduct",
        )
        replay_deferred_receipts()

        self.assertFalse(DeferredReceipt.objects.exists())
        self.assertEqual(
            ReceiptSaleLine.objects.get(receipt_id=self.receipt_id, product=product).quantity,
            Decimal("6"),
        )
        self.assertEqual(
            SalesDailySummary.objects.get(store=self.store, product=product).quantity_sold,
            Decimal("6"),
        )

    def test_cancelled_receipt_removes_previous_contribution(self):
        affected, counts = set(), {"created": 0, "updated": 0}
        _apply_receipt(self.receipt(3), affected, counts)
        _apply_receipt(self.receipt(3, revision=2, cancelled=True), affected, counts)
        self.assertFalse(SalesDailySummary.objects.filter(store=self.store, product=self.product).exists())
        self.assertFalse(ReceiptSaleLine.objects.exists())

    def test_return_offsets_sales_regardless_of_import_order(self):
        returned = self.receipt(-2, revision=1)
        sold = self.receipt(5, revision=2)
        sold["id"] = str(uuid.uuid4())

        affected, counts = set(), {"created": 0, "updated": 0}
        _apply_receipt(returned, affected, counts)
        _apply_receipt(sold, affected, counts)

        summary = SalesDailySummary.objects.get(store=self.store, product=self.product)
        self.assertEqual(summary.quantity_sold, Decimal("3"))
        self.assertEqual(summary.receipts_count, 2)

    def test_supplier_names_use_compact_catalog_aliases(self):
        self.assertEqual(supplier_short_name("Southern Glazer's Wine & Spirits of FL"), "Southern")
        self.assertEqual(supplier_short_name("Republic National Dist - Tampa"), "RNDC")
        self.assertEqual(supplier_short_name("Green Light Distribution Florida"), "GLDF")

    def test_older_backfill_revision_cannot_replace_newer_receipt(self):
        affected, counts = set(), {"created": 0, "updated": 0}
        _apply_receipt(self.receipt(5, revision=10), affected, counts)
        _apply_receipt(self.receipt(2, revision=9), affected, counts)
        summary = SalesDailySummary.objects.get(store=self.store, product=self.product)
        line = ReceiptSaleLine.objects.get()
        self.assertEqual(summary.quantity_sold, Decimal("5"))
        self.assertEqual(line.quantity, Decimal("5"))
        self.assertEqual(line.receipt_revision, 10)

    @patch("orders.services.KoronaClient")
    def test_incremental_sync_only_requests_live_revisions(self, client_class):
        live = SyncState.objects.create(entity="receipts_live", last_revision=120)
        recent = SyncState.objects.create(
            entity="receipts_recent", cursor_data={"complete": False, "next_date": "2026-05-18"}
        )
        historical = SyncState.objects.create(entity="receipts", last_revision=14)
        client = client_class.return_value
        client.account_path.return_value = "accounts/test/receipts"
        client.request.return_value = {"results": [], "maxRevision": 125}

        counts = sync_receipts()

        client.request.assert_called_once()
        params = client.request.call_args.kwargs["params"]
        self.assertEqual(params["revision"], 120)
        self.assertEqual(params["sort"], "revision")
        self.assertNotIn("maxBookingTime", params)
        live.refresh_from_db()
        recent.refresh_from_db()
        historical.refresh_from_db()
        self.assertEqual(live.last_revision, 125)
        self.assertEqual(recent.cursor_data["next_date"], "2026-05-18")
        self.assertEqual(historical.last_revision, 14)
        self.assertEqual(counts["starting_revision"], 120)
        self.assertEqual(counts["ending_revision"], 125)
        self.assertEqual(counts["products_recalculated"], 0)

    @patch("orders.services.KoronaClient")
    def test_incremental_sync_updates_changed_receipt_and_affected_total(self, client_class):
        affected, initial_counts = set(), {"created": 0, "updated": 0}
        _apply_receipt(self.receipt(2, revision=10), affected, initial_counts)
        SyncState.objects.create(entity="receipts_live", last_revision=10)
        changed = self.receipt(5, revision=11)
        client = client_class.return_value
        client.account_path.return_value = "accounts/test/receipts"
        client.request.return_value = {"results": [changed], "maxRevision": 11}

        counts = sync_receipts()

        line = ReceiptSaleLine.objects.get(receipt_id=self.receipt_id, product=self.product)
        total = self.product.productmonthlyneed_set.get(store=self.store)
        self.assertEqual(line.quantity, Decimal("5"))
        self.assertEqual(total.needed_quantity, Decimal("5"))
        self.assertEqual(counts["seen"], 1)
        self.assertEqual(counts["updated"], 1)
        self.assertEqual(counts["products_recalculated"], 1)

    def test_changed_monthly_totals_are_recalculated_in_bulk(self):
        second_product = Product.objects.create(
            korona_id=uuid.uuid4(), number="200", name="Second Bottle"
        )
        today = timezone.localdate()
        SalesDailySummary.objects.create(
            store=self.store, product=self.product, sales_date=today, quantity_sold=3
        )
        SalesDailySummary.objects.create(
            store=self.store, product=second_product, sales_date=today, quantity_sold=7
        )

        updated = recalculate_monthly_needs_for_pairs(
            {(self.store.id, self.product.id), (self.store.id, second_product.id)}
        )

        self.assertEqual(updated, 2)
        self.assertEqual(
            self.product.productmonthlyneed_set.get(store=self.store).needed_quantity,
            Decimal("3"),
        )
        self.assertEqual(
            second_product.productmonthlyneed_set.get(store=self.store).needed_quantity,
            Decimal("7"),
        )

    def test_monthly_need_is_exact_trailing_30_day_sales(self):
        today = timezone.localdate()
        SalesDailySummary.objects.create(
            store=self.store, product=self.product, sales_date=today, quantity_sold=Decimal("12.5")
        )
        SalesDailySummary.objects.create(
            store=self.store, product=self.product, sales_date=today - timedelta(days=29), quantity_sold=4
        )
        SalesDailySummary.objects.create(
            store=self.store, product=self.product, sales_date=today - timedelta(days=30), quantity_sold=100
        )
        SalesDailySummary.objects.create(
            store=self.store, product=self.product, sales_date=today + timedelta(days=1), quantity_sold=200
        )

        total = recalculate_monthly_need(self.store.id, self.product.id, today.replace(day=1))

        self.assertEqual(total.needed_quantity, Decimal("16.5"))

    def test_stale_monthly_totals_are_refreshed_in_bulk(self):
        today = timezone.localdate()
        SalesDailySummary.objects.create(
            store=self.store, product=self.product, sales_date=today, quantity_sold=7
        )
        cached = self.product.productmonthlyneed_set.create(
            store=self.store,
            month=today.replace(day=1),
            needed_quantity=999,
            last_calculated_at=timezone.now() - timedelta(days=1),
        )

        self.assertEqual(recalculate_stale_monthly_needs(), 1)
        cached.refresh_from_db()
        self.assertEqual(cached.needed_quantity, 7)

    def test_full_monthly_rebuild_replaces_all_current_totals(self):
        today = timezone.localdate()
        ReceiptSaleLine.objects.create(
            receipt_id=uuid.uuid4(),
            receipt_revision=10,
            store=self.store,
            product=self.product,
            sales_date=today,
            quantity=9,
        )
        self.product.productmonthlyneed_set.create(
            store=self.store, month=today.replace(day=1), needed_quantity=999
        )

        counts = rebuild_all_monthly_needs()

        rebuilt = self.product.productmonthlyneed_set.get(store=self.store, month=today.replace(day=1))
        self.assertEqual(rebuilt.needed_quantity, 9)
        self.assertEqual(rebuilt.avg_daily_sales_30, Decimal("0.3"))
        self.assertEqual(counts["monthly_totals_rebuilt"], 1)
        self.assertEqual(counts["daily_summaries_rebuilt"], 1)

    @patch("orders.services.KoronaClient")
    def test_full_month_download_replaces_stale_day_before_rebuilding(self, client_class):
        today = timezone.localdate()
        stale_receipt_id = uuid.uuid4()
        ReceiptSaleLine.objects.create(
            receipt_id=stale_receipt_id,
            receipt_revision=1,
            store=self.store,
            product=self.product,
            sales_date=today,
            quantity=99,
        )
        SalesDailySummary.objects.create(
            store=self.store,
            product=self.product,
            sales_date=today,
            quantity_sold=99,
            receipts_count=1,
            last_receipt_revision=1,
        )
        fresh_receipt = self.receipt(4, revision=8)
        fresh_receipt["bookingTime"] = timezone.now().isoformat()
        client = client_class.return_value
        client.account_path.return_value = "accounts/test/receipts"
        client.request.return_value = {"results": [fresh_receipt]}

        counts = sync_monthly_receipt_history(days=1)

        self.assertFalse(ReceiptSaleLine.objects.filter(receipt_id=stale_receipt_id).exists())
        fresh_line = ReceiptSaleLine.objects.get(receipt_id=self.receipt_id)
        self.assertEqual(fresh_line.quantity, Decimal("4"))
        summary = SalesDailySummary.objects.get(
            store=self.store, product=self.product, sales_date=today
        )
        self.assertEqual(summary.quantity_sold, Decimal("4"))
        self.assertEqual(counts["seen"], 1)
        params = client.request.call_args.kwargs["params"]
        self.assertEqual(params["voidedItems"], "true")
        self.assertEqual(params["page"], 1)
        self.assertIn("minBookingTime", params)
        self.assertIn("maxBookingTime", params)

    def test_nightly_monthly_reconciliation_fetches_full_window_before_rebuilding(self):
        with patch(
            "orders.services.sync_monthly_receipt_history",
            return_value={"seen": 5, "created": 1, "updated": 2, "deferred": 0, "days_refreshed": 30},
        ) as replay, patch(
            "orders.services.rebuild_all_monthly_needs",
            return_value={"seen": 3, "created": 3, "updated": 0, "daily_summaries_rebuilt": 9, "monthly_totals_rebuilt": 3},
        ) as rebuild:
            counts = reconcile_monthly_needs()

        replay.assert_called_once_with(days=30)
        rebuild.assert_called_once_with()
        self.assertEqual(counts["seen"], 5)
        self.assertEqual(counts["monthly_totals_rebuilt"], 3)
        self.assertEqual(counts["daily_summaries_rebuilt"], 9)
        self.assertEqual(counts["days_refreshed"], 30)


class OrderApiTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user("buyer", password="secret")
        self.client.force_login(self.user)
        self.store = Store.objects.create(korona_id=uuid.uuid4(), number="1", name="Main")

    def test_create_order_list_is_idempotent(self):
        payload = {"store_id": self.store.id, "order_date": "2026-07-11"}
        first = self.client.post("/api/orders/", payload, content_type="application/json")
        second = self.client.post("/api/orders/", payload, content_type="application/json")
        self.assertEqual(first.status_code, 201)
        self.assertEqual(second.status_code, 200)
        self.assertEqual(OrderList.objects.count(), 1)

    def test_product_suggestions_include_current_store_stock(self):
        product = Product.objects.create(
            korona_id=uuid.uuid4(),
            number="1027",
            name="BUCHANAN'S",
            normalized_name="buchanans",
            stock_last_synced_at=timezone.now(),
        )
        ProductStock.objects.create(product=product, store=self.store, actual=7)
        order_list = OrderList.objects.create(
            store=self.store, order_date=date(2026, 7, 12), created_by=self.user
        )
        response = self.client.get(f"/api/products/search/?q=buchanans&order_id={order_list.id}")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()[0]["current_stock"], 7.0)

    def test_order_list_includes_supplier_fields_and_admin_can_choose_preferred(self):
        first_id = uuid.uuid4()
        second_id = uuid.uuid4()
        product = Product.objects.create(
            korona_id=uuid.uuid4(),
            number="SUP-1",
            name="Supplier Test",
            normalized_name="suppliertest",
            raw_data={
                "commodityGroup": {"name": "Whiskey", "number": "17"},
                "supplierPrices": [
                    {
                        "supplier": {"id": str(first_id), "name": "Alpha Supply", "number": "10"},
                        "orderCode": "A-100",
                        "containerSize": 6,
                        "value": 12.5,
                    },
                    {
                        "supplier": {"id": str(second_id), "name": "Beta Supply", "number": "20"},
                        "orderCode": "B-200",
                        "containerSize": 12,
                        "value": 22.75,
                    },
                ]
            },
        )
        order_list = OrderList.objects.create(
            store=self.store, order_date=date(2026, 7, 25), created_by=self.user
        )
        OrderListItem.objects.create(
            order_list=order_list, product=product, created_by=self.user, updated_by=self.user
        )

        initial = self.client.get(f"/api/orders/{order_list.id}/").json()["items"][0]
        self.assertEqual(initial["supplier_name"], "Alpha Supply")
        self.assertEqual(initial["supplier_full_name"], "Alpha Supply")
        self.assertEqual(initial["supplier_order_code"], "A-100")
        self.assertEqual(initial["supplier_pack_size"], 6.0)
        self.assertEqual(initial["supplier_purchase_price"], 12.5)
        self.assertEqual(initial["commodity_group"], "Whiskey")
        self.assertEqual(initial["commodity_group_number"], "17")

        self.assertEqual(
            self.client.patch(
                f"/api/products/{product.id}/preferred-supplier/",
                {"supplier_id": str(second_id)},
                content_type="application/json",
            ).status_code,
            403,
        )
        self.user.is_staff = True
        self.user.save(update_fields=["is_staff"])
        selected = self.client.patch(
            f"/api/products/{product.id}/preferred-supplier/",
            {"supplier_id": str(second_id)},
            content_type="application/json",
        )
        self.assertEqual(selected.status_code, 200)
        self.assertEqual(selected.json()["supplier_name"], "Beta Supply")
        self.assertEqual(selected.json()["supplier_number"], "20")
        self.assertEqual(selected.json()["supplier_order_code"], "B-200")
        self.assertEqual(selected.json()["supplier_pack_size"], 12.0)
        self.assertEqual(selected.json()["supplier_purchase_price"], 22.75)
        product.refresh_from_db()
        self.assertEqual(product.preferred_supplier_id, second_id)

    def test_exact_product_number_ranks_before_fuzzy_matches(self):
        exact = Product.objects.create(
            korona_id=uuid.uuid4(),
            number="10001",
            name="Exact Product",
            normalized_name="exactproduct",
            stock_last_synced_at=timezone.now(),
        )
        Product.objects.create(
            korona_id=uuid.uuid4(),
            number="99999",
            name="$10,001 Holiday Cash",
            normalized_name="10001holidaycash",
            stock_last_synced_at=timezone.now(),
        )
        order_list = OrderList.objects.create(
            store=self.store, order_date=date(2026, 7, 15), created_by=self.user
        )
        response = self.client.get(f"/api/products/search/?q=10001&order_id={order_list.id}")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()[0]["id"], exact.id)

    def test_barcode_search_tolerates_one_leading_or_trailing_zero(self):
        product = Product.objects.create(
            korona_id=uuid.uuid4(), number="BC-1", name="Barcode Product", normalized_name="barcodeproduct"
        )
        ProductCode.objects.create(product=product, code="123456789012", normalized_code="123456789012")

        for scanned in ["0123456789012", "1234567890120", "123456789012"]:
            response = self.client.get(f"/api/products/search/?q={scanned}")
            self.assertEqual(response.status_code, 200)
            self.assertEqual([row["id"] for row in response.json()], [product.id])

    def test_product_search_returns_all_literal_matches_without_fuzzy_noise(self):
        for index in range(25):
            Product.objects.create(
                korona_id=uuid.uuid4(),
                number=f"4PK-{index}",
                name=f"Canned Cocktail 4PK Flavor {index}",
                normalized_name=f"cannedcocktail4pkflavor{index}",
                stock_last_synced_at=timezone.now(),
            )
        Product.objects.create(
            korona_id=uuid.uuid4(),
            number="RYAN-1",
            name="RYAN'S 1.75L",
            normalized_name="ryans175l",
            stock_last_synced_at=timezone.now(),
        )
        rye = Product.objects.create(
            korona_id=uuid.uuid4(),
            number="RYE-1",
            name="STRAIGHT RYE 750ML",
            normalized_name="straightrye750ml",
            stock_last_synced_at=timezone.now(),
        )

        four_packs = self.client.get("/api/products/search/?q=4pk")
        rye_results = self.client.get("/api/products/search/?q=rye")

        self.assertEqual(len(four_packs.json()), 25)
        self.assertEqual([row["id"] for row in rye_results.json()], [rye.id])

    def test_numeric_name_search_does_not_partially_match_product_number(self):
        ninety_nine = Product.objects.create(
            korona_id=uuid.uuid4(),
            number="3330",
            name="99 50ML VODKA",
            normalized_name="9950mlvodka",
        )
        Product.objects.create(
            korona_id=uuid.uuid4(),
            number="3499",
            name="SMIRNOFF 50ML RASPBERRY",
            normalized_name="smirnoff50mlraspberry",
        )

        response = self.client.get("/api/products/search/?q=99")

        self.assertEqual([row["id"] for row in response.json()], [ninety_nine.id])

    def test_multi_word_search_returns_and_ranks_all_matching_sizes(self):
        expected = []
        for size in ["50ML", "200ML", "375ML", "750ML"]:
            product = Product.objects.create(
                korona_id=uuid.uuid4(),
                number=f"BAN-{size}",
                name=f"99 {size} BANANA",
                normalized_name=f"99{size.lower()}banana",
            )
            expected.append(product.id)
        Product.objects.create(
            korona_id=uuid.uuid4(),
            number="OTHER-99",
            name="99 APPLES 375ML",
            normalized_name="99apples375ml",
        )
        Product.objects.create(
            korona_id=uuid.uuid4(),
            number="OTHER-BAN",
            name="BANANAS RUM 750ML",
            normalized_name="bananasrum750ml",
        )

        response = self.client.get("/api/products/search/?q=99%20bananas")

        self.assertEqual(response.status_code, 200)
        self.assertEqual({row["id"] for row in response.json()}, set(expected))

        order_list = OrderList.objects.create(
            store=self.store, order_date=date(2026, 7, 20), created_by=self.user
        )
        main_order = self.client.post(
            f"/api/orders/{order_list.id}/items/bulk/",
            {"product_ids": [expected[0]]},
            content_type="application/json",
        )
        inventory = self.client.post(
            "/api/inventory/compare/",
            {"product_ids": [expected[1]]},
            content_type="application/json",
        )
        self.user.is_staff = True
        self.user.save(update_fields=["is_staff"])
        bulk_order = BulkOrderList.objects.create(name="Search test", created_by=self.user)
        bulk = self.client.post(
            f"/api/bulk-orders/{bulk_order.id}/items/",
            {"product_ids": [expected[2]]},
            content_type="application/json",
        )

        self.assertEqual(main_order.status_code, 201)
        self.assertEqual([row["id"] for row in inventory.json()["products"]], [expected[1]])
        self.assertEqual(bulk.status_code, 200)
        self.assertTrue(bulk_order.items.filter(product_id=expected[2]).exists())

    def test_inactive_listed_products_are_hidden_and_unavailable_for_new_orders(self):
        inactive = Product.objects.create(
            korona_id=uuid.uuid4(),
            number="LEGACY-TITO",
            name="TITO'S 50ML",
            normalized_name="titos50ml",
            active=False,
            raw_data={"listed": True, "deactivated": False},
        )
        active = Product.objects.create(
            korona_id=uuid.uuid4(),
            number="ACTIVE-TITO",
            name="TITO'S 50ML",
            normalized_name="titos50ml",
            active=True,
        )
        order_list = OrderList.objects.create(
            store=self.store, order_date=date(2026, 7, 21), created_by=self.user
        )

        search = self.client.get("/api/products/search/?q=tito")
        add_inactive = self.client.post(
            f"/api/orders/{order_list.id}/items/bulk/",
            {"product_ids": [inactive.id]},
            content_type="application/json",
        )

        self.assertEqual([row["id"] for row in search.json()], [active.id])
        self.assertEqual(add_inactive.status_code, 400)
        self.assertFalse(order_list.items.filter(product=inactive).exists())

        self.user.is_staff = True
        self.user.is_superuser = True
        self.user.save(update_fields=["is_staff", "is_superuser"])
        enabled = self.client.patch(
            "/api/operations/services/",
            {"show_inactive_products": True},
            content_type="application/json",
        )
        visible_search = self.client.get("/api/products/search/?q=tito")
        add_visible_inactive = self.client.post(
            f"/api/orders/{order_list.id}/items/bulk/",
            {"product_ids": [inactive.id]},
            content_type="application/json",
        )

        self.assertEqual(enabled.status_code, 200)
        self.assertTrue(SystemSetting.objects.get(key="show_inactive_products").value)
        self.assertEqual(
            {row["id"] for row in visible_search.json()}, {active.id, inactive.id}
        )
        self.assertEqual(add_visible_inactive.status_code, 201)
        self.assertTrue(order_list.items.filter(product=inactive).exists())

    def test_multi_word_search_matches_words_separated_in_catalog_name(self):
        product = Product.objects.create(
            korona_id=uuid.uuid4(),
            number="BAN-BRAND",
            name="99 BRAND BANANA 375 ML",
            normalized_name="99brandbanana375ml",
        )

        response = self.client.get("/api/products/search/?q=99%20bananas%20375ml")

        self.assertEqual([row["id"] for row in response.json()], [product.id])

    def test_category_keyword_overrides_are_saved(self):
        response = self.client.put(
            "/api/product-categories/",
            {"hidden": [], "custom": [], "overrides": {"vodka": "vodka,neutral spirit"}},
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["overrides"]["vodka"], "vodka,neutral spirit")
        self.assertEqual(self.client.get("/api/product-categories/").json()["overrides"]["vodka"], "vodka,neutral spirit")

    def test_product_search_corrects_pinot_nior_typo(self):
        pinot = Product.objects.create(
            korona_id=uuid.uuid4(),
            number="PN-1",
            name="HOUSE PINOT NOIR 750ML",
            normalized_name="housepinotnoir750ml",
            stock_last_synced_at=timezone.now(),
        )

        response = self.client.get("/api/products/search/?q=pinot%20nior")

        self.assertEqual([row["id"] for row in response.json()], [pinot.id])

    def test_bulk_add_products_preserves_existing_item_values(self):
        order_list = OrderList.objects.create(
            store=self.store, order_date=date(2026, 7, 19), created_by=self.user
        )
        products = [
            Product.objects.create(
                korona_id=uuid.uuid4(),
                number=str(900 + index),
                name=f"Batch Product {index}",
                normalized_name=f"batchproduct{index}",
            )
            for index in range(3)
        ]
        existing = OrderListItem.objects.create(
            order_list=order_list,
            product=products[0],
            on_shelf_quantity=7,
            joe_quantity=3,
            row_order=4,
            created_by=self.user,
            updated_by=self.user,
        )

        response = self.client.post(
            f"/api/orders/{order_list.id}/items/bulk/",
            {"product_ids": [products[0].id, products[1].id, products[2].id, products[1].id]},
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 201)
        self.assertEqual((response.json()["created"], response.json()["existing"]), (2, 1))
        existing.refresh_from_db()
        self.assertEqual(existing.on_shelf_quantity, 7)
        self.assertEqual(existing.joe_quantity, 3)
        added = list(order_list.items.exclude(pk=existing.pk).order_by("row_order"))
        self.assertEqual([item.product_id for item in added], [products[1].id, products[2].id])
        self.assertEqual([item.row_order for item in added], [0, 1])
        self.assertEqual([item.on_shelf_quantity for item in added], [0, 0])

    def test_pdf_and_xlsx_exports(self):
        self.user.is_staff = True
        self.user.save(update_fields=["is_staff"])
        order_list = OrderList.objects.create(
            store=self.store, order_date=date(2026, 7, 13), created_by=self.user
        )
        product = Product.objects.create(
            korona_id=uuid.uuid4(), number="55", name="Test Spirit", normalized_name="testspirit"
        )
        ProductCode.objects.create(product=product, code="012345678901", normalized_code="012345678901")
        item = OrderListItem.objects.create(
            order_list=order_list,
            product=product,
            joe_quantity=2,
            bt_quantity=3,
            sqw_quantity=4,
            created_by=self.user,
            updated_by=self.user,
        )
        other_store = Store.objects.create(korona_id=uuid.uuid4(), number="2", name="Other")
        OrderItemTransfer.objects.create(item=item, from_store=other_store, quantity=5)
        cases = []
        for kind in ["joe", "bt", "sqw", "transfers"]:
            cases.extend(
                [
                    (f"/api/orders/{order_list.id}/export/{kind}.pdf", "application/pdf"),
                    (
                        f"/api/orders/{order_list.id}/export/{kind}.xlsx",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    ),
                ]
            )
        for url, content_type in cases:
            response = self.client.get(url)
            self.assertEqual(response.status_code, 200)
            self.assertEqual(response["Content-Type"], content_type)
        self.assertEqual(self.client.get(f"/api/orders/{order_list.id}/export/order.pdf").status_code, 404)
        self.assertEqual(self.client.get(f"/api/orders/{order_list.id}/export/order.xlsx").status_code, 404)

        joe_response = self.client.get(f"/api/orders/{order_list.id}/export/joe.xlsx")
        joe_sheet = load_workbook(BytesIO(joe_response.content)).active
        self.assertEqual([cell.value for cell in joe_sheet[4]], ["Product Name", "Product #", "Barcode", "JOE"])
        self.assertEqual([cell.value for cell in joe_sheet[5]], ["Test Spirit", "55", "012345678901", 2])

        grid_payload = {
            "orientation": "landscape",
            "title": "Monday Transfers",
            "columns": [
                {"id": "product_number", "label": "Product #", "width": 110},
                {"id": "product_name", "label": "Product name", "width": 210},
                {"id": "supplier_name", "label": "Supplier", "width": 120},
            ],
            "rows": [["55", "Test Spirit", "Southern"], ["56", "Filtered & sorted", "RNDC"]],
        }
        grid_response = self.client.post(
            f"/api/orders/{order_list.id}/export-grid.pdf",
            grid_payload,
            content_type="application/json",
            HTTP_ACCEPT="*/*",
        )
        self.assertEqual(grid_response.status_code, 200)
        self.assertEqual(grid_response["Content-Type"], "application/pdf")
        self.assertTrue(grid_response.content.startswith(b"%PDF"))
        self.assertIn("landscape", grid_response["X-PDF-Filename"])
        self.assertIn("Monday_Transfers", grid_response["X-PDF-Filename"])

        grid_xlsx_response = self.client.post(
            f"/api/orders/{order_list.id}/export-grid.xlsx",
            grid_payload,
            content_type="application/json",
            HTTP_ACCEPT="*/*",
        )
        self.assertEqual(grid_xlsx_response.status_code, 200)
        self.assertEqual(
            grid_xlsx_response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        grid_sheet = load_workbook(BytesIO(grid_xlsx_response.content)).active
        self.assertEqual(grid_sheet["A1"].value, "Monday Transfers")
        self.assertEqual([cell.value for cell in grid_sheet[4]], ["Product #", "Product name", "Supplier"])
        self.assertEqual([cell.value for cell in grid_sheet[5]], ["55", "Test Spirit", "Southern"])

        grid_payload["orientation"] = "portrait"
        portrait_response = self.client.post(
            f"/api/orders/{order_list.id}/export-grid.pdf",
            grid_payload,
            content_type="application/json",
        )
        self.assertEqual(portrait_response.status_code, 200)
        self.assertIn("portrait", portrait_response["X-PDF-Filename"])

    def test_transfer_export_sorts_and_colors_rows_by_store(self):
        self.user.is_staff = True
        self.user.save(update_fields=["is_staff"])
        order_list = OrderList.objects.create(
            store=self.store, order_date=date(2026, 7, 17), created_by=self.user
        )
        product = Product.objects.create(
            korona_id=uuid.uuid4(), number="500", name="Grouped Transfer", normalized_name="groupedtransfer"
        )
        item = OrderListItem.objects.create(
            order_list=order_list, product=product, created_by=self.user, updated_by=self.user
        )
        store_20 = Store.objects.create(korona_id=uuid.uuid4(), number="20", name="Twenty")
        store_3 = Store.objects.create(korona_id=uuid.uuid4(), number="3", name="Three")
        OrderItemTransfer.objects.create(item=item, from_store=store_20, quantity=2)
        OrderItemTransfer.objects.create(item=item, from_store=store_3, quantity=3)

        response = self.client.get(f"/api/orders/{order_list.id}/export/transfers.xlsx")
        sheet = load_workbook(BytesIO(response.content)).active
        self.assertEqual([sheet.cell(row=row, column=3).value for row in (5, 6)], ["#3", "#20"])
        self.assertNotEqual(sheet["A5"].fill.fgColor.rgb, sheet["A6"].fill.fgColor.rgb)

    def test_order_item_uses_only_distributor_quantities(self):
        self.user.is_staff = True
        self.user.save(update_fields=["is_staff"])
        order_list = OrderList.objects.create(
            store=self.store, order_date=date(2026, 7, 14), created_by=self.user
        )
        product = Product.objects.create(
            korona_id=uuid.uuid4(), number="200", name="Distributor Item", normalized_name="distributoritem"
        )
        item = OrderListItem.objects.create(
            order_list=order_list, product=product, created_by=self.user, updated_by=self.user
        )
        response = self.client.patch(
            f"/api/items/{item.id}/",
            {"joe_quantity": 1, "bt_quantity": 2, "sqw_quantity": 3},
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual((payload["joe_quantity"], payload["bt_quantity"], payload["sqw_quantity"]), ("1.000", "2.000", "3.000"))
        self.assertNotIn("final_quantity", payload)
        self.assertNotIn("suggested_quantity", payload)

    def test_order_item_accepts_multiple_transfer_stores(self):
        self.user.is_staff = True
        self.user.save(update_fields=["is_staff"])
        order_list = OrderList.objects.create(
            store=self.store, order_date=date(2026, 7, 16), created_by=self.user
        )
        product = Product.objects.create(
            korona_id=uuid.uuid4(), number="300", name="Transfer Item", normalized_name="transferitem"
        )
        item = OrderListItem.objects.create(
            order_list=order_list, product=product, created_by=self.user, updated_by=self.user
        )
        first = Store.objects.create(korona_id=uuid.uuid4(), number="2", name="Second")
        second = Store.objects.create(korona_id=uuid.uuid4(), number="3", name="Third")
        response = self.client.patch(
            f"/api/items/{item.id}/",
            {
                "transfers": [
                    {"from_store_id": first.id, "quantity": 4},
                    {"from_store_id": second.id, "quantity": 7},
                ]
            },
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(OrderItemTransfer.objects.filter(item=item).count(), 2)
        self.assertEqual(
            {(row["from_store_number"], row["quantity"]) for row in response.json()["transfers"]},
            {("2", 4.0), ("3", 7.0)},
        )

    @override_settings(
        STORAGES={"staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"}}
    )
    def test_normal_user_cannot_view_operations_or_run_jobs(self):
        self.assertEqual(self.client.get("/ops/").status_code, 403)
        response = self.client.get("/api/operations/services/")
        self.assertEqual(response.status_code, 403)
        self.assertEqual(self.client.post("/api/operations/services/stores/run/").status_code, 403)

    def test_operations_api_reports_per_store_stock_cursor_health(self):
        self.user.is_staff = True
        self.user.is_superuser = True
        self.user.save(update_fields=["is_staff", "is_superuser"])
        self.store.is_warehouse = True
        self.store.save(update_fields=["is_warehouse"])
        product = Product.objects.create(
            korona_id=uuid.uuid4(), number="OPS-1", name="Ops Stock", normalized_name="opsstock"
        )
        ProductStock.objects.create(product=product, store=self.store, actual=4)
        SyncState.objects.create(
            entity=STOCK_SYNC_ENTITY,
            store=self.store,
            last_revision=456,
            last_synced_at=timezone.now(),
        )
        SyncRun.objects.create(
            job_name="stocks",
            status=SyncRun.Status.SUCCESS,
            finished_at=timezone.now(),
            duration_ms=1200,
            records_seen=3,
            records_updated=3,
            metrics={"stores_checked": 1, "seen": 3, "created": 0, "updated": 3, "unchanged": 0},
        )

        response = self.client.get("/api/operations/services/")

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["counts"]["active_stores"], 1)
        self.assertEqual(payload["stock_sync"]["health"], "healthy")
        self.assertEqual(payload["stock_sync"]["current"], 1)
        self.assertEqual(payload["stock_sync"]["latest_incremental"]["changed"], 3)
        self.assertEqual(payload["stock_sync"]["stores"][0]["last_revision"], 456)
        self.assertEqual(payload["stock_sync"]["stores"][0]["stock_records"], 1)
        self.assertEqual(payload["overview"]["records_changed_24h"], 3)
        stock_service = next(service for service in payload["services"] if service["name"] == "stocks")
        self.assertEqual(stock_service["latest_run"]["metrics"]["stores_checked"], 1)
        self.assertIn("description", payload["services"][0])
        self.assertEqual(
            {service["name"] for service in payload["services"]},
            {"stores", "products", "stocks", "stock_reconciliation", "receipts", "monthly_reconciliation"},
        )

    @patch("orders.tasks.sync_stores_task.delay")
    def test_manual_job_is_marked_queued_and_cannot_be_duplicated(self, delay):
        delay.return_value.id = "queued-task"
        self.user.is_staff = True
        self.user.is_superuser = True
        self.user.save(update_fields=["is_staff", "is_superuser"])

        first = self.client.post("/api/operations/services/stores/run/")
        second = self.client.post("/api/operations/services/stores/run/")

        self.assertEqual(first.status_code, 202)
        self.assertEqual(second.status_code, 409)
        control = ServiceControl.objects.get(service_name="stores")
        self.assertEqual(control.status, ServiceControl.Status.QUEUED)
        delay.assert_called_once_with(force=True)

    def test_operations_rejects_invalid_or_fixed_schedule_intervals(self):
        self.user.is_staff = True
        self.user.is_superuser = True
        self.user.save(update_fields=["is_staff", "is_superuser"])

        invalid = self.client.patch(
            "/api/operations/services/",
            {"service_name": "stocks", "interval_seconds": "fast"},
            content_type="application/json",
        )
        fixed = self.client.patch(
            "/api/operations/services/",
            {"service_name": "stock_reconciliation", "interval_seconds": 300},
            content_type="application/json",
        )

        self.assertEqual(invalid.status_code, 400)
        self.assertEqual(fixed.status_code, 400)

    def test_operations_controls_can_be_disabled_restored_and_rescheduled(self):
        self.user.is_staff = True
        self.user.is_superuser = True
        self.user.save(update_fields=["is_staff", "is_superuser"])

        disabled = self.client.patch(
            "/api/operations/services/",
            {"service_name": "stores", "enabled": False},
            content_type="application/json",
        )
        restored = self.client.patch(
            "/api/operations/services/",
            {"service_name": "stores", "enabled": True, "interval_seconds": 600},
            content_type="application/json",
        )

        self.assertEqual(disabled.status_code, 200)
        self.assertEqual(restored.status_code, 200)
        control = ServiceControl.objects.get(service_name="stores")
        self.assertTrue(control.enabled)
        self.assertEqual(control.status, ServiceControl.Status.IDLE)
        self.assertEqual(control.interval_seconds, 600)
        self.assertEqual(control.updated_by, self.user)

    def test_every_operations_run_button_queues_the_correct_task(self):
        self.user.is_staff = True
        self.user.is_superuser = True
        self.user.save(update_fields=["is_staff", "is_superuser"])
        task_targets = {
            "stores": "orders.tasks.sync_stores_task.delay",
            "products": "orders.tasks.sync_products_task.delay",
            "stocks": "orders.tasks.sync_stocks_task.delay",
            "stock_reconciliation": "orders.tasks.reconcile_stocks_task.delay",
            "receipts": "orders.tasks.sync_receipts_task.delay",
            "monthly_reconciliation": "orders.tasks.reconcile_monthly_totals_task.delay",
        }

        for service_name, target in task_targets.items():
            with self.subTest(service_name=service_name), patch(target) as delay:
                delay.return_value.id = f"{service_name}-task"
                response = self.client.post(f"/api/operations/services/{service_name}/run/")
                self.assertEqual(response.status_code, 202)
                self.assertEqual(response.json()["task_id"], f"{service_name}-task")
                delay.assert_called_once_with(force=True)

    @patch("orders.tasks.sync_stores_task.delay", side_effect=RuntimeError("broker unavailable"))
    def test_operations_queue_failure_is_visible_and_recoverable(self, delay):
        self.user.is_staff = True
        self.user.is_superuser = True
        self.user.save(update_fields=["is_staff", "is_superuser"])

        response = self.client.post("/api/operations/services/stores/run/")

        self.assertEqual(response.status_code, 503)
        control = ServiceControl.objects.get(service_name="stores")
        self.assertEqual(control.status, ServiceControl.Status.ERROR)
        self.assertIn("broker unavailable", control.last_error)

    def test_operations_diagnostics_report_errors_latency_endpoints_and_context(self):
        self.user.is_staff = True
        self.user.is_superuser = True
        self.user.save(update_fields=["is_staff", "is_superuser"])
        ServiceControl.objects.create(
            service_name="products",
            status=ServiceControl.Status.ERROR,
            last_error="KORONA timed out",
        )
        ApiRequestLog.objects.create(
            method="GET", url_path="/products", status_code=200, latency_ms=200
        )
        ApiRequestLog.objects.create(
            method="GET", url_path="/products", status_code=503, latency_ms=1400
        )
        SystemLog.objects.create(
            level="ERROR",
            source="products",
            message="Product import failed",
            context={"page": 17, "retryable": True},
        )

        response = self.client.get("/api/operations/services/")

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["overview"]["errors"], 1)
        self.assertEqual(payload["overview"]["attention"][0]["name"], "products")
        self.assertEqual(payload["api_health"]["requests_24h"], 2)
        self.assertEqual(payload["api_health"]["errors_24h"], 1)
        self.assertEqual(payload["api_health"]["average_ms"], 800)
        self.assertEqual(payload["api_health"]["slow_requests_24h"], 1)
        self.assertEqual(payload["api_health"]["endpoints"][0]["errors"], 1)
        self.assertEqual(payload["api_latency"][0]["method"], "GET")
        self.assertEqual(payload["logs"][0]["context"]["page"], 17)

    @override_settings(
        STORAGES={"staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"}}
    )
    def test_operations_page_contains_all_diagnostic_controls_and_no_full_resync(self):
        self.user.is_staff = True
        self.user.is_superuser = True
        self.user.save(update_fields=["is_staff", "is_superuser"])

        response = self.client.get("/ops/")

        self.assertEqual(response.status_code, 200)
        body = response.content.decode()
        for element_id in (
            "ops-attention",
            "show-inactive-products",
            "stock-coverage",
            "run-service-filter",
            "run-status-filter",
            "log-level-filter",
            "log-search",
            "latency-status-filter",
            "latency-search",
            "endpoint-health-list",
        ):
            self.assertIn(f'id="{element_id}"', body)
        self.assertNotIn("Run full resync", body)
        self.assertEqual(self.client.post("/api/operations/services/full_sync/run/").status_code, 404)

    def test_normal_user_gets_only_the_restricted_working_columns(self):
        order_list = OrderList.objects.create(
            store=self.store, order_date=date(2026, 7, 20), created_by=self.user
        )
        product = Product.objects.create(
            korona_id=uuid.uuid4(), number="901", name="Restricted", normalized_name="restricted"
        )
        item = OrderListItem.objects.create(
            order_list=order_list,
            product=product,
            joe_quantity=8,
            created_by=self.user,
            updated_by=self.user,
        )

        payload = self.client.get(f"/api/orders/{order_list.id}/").json()

        self.assertEqual(payload["stores"], [])
        self.assertFalse(payload["permissions"]["is_admin"])
        self.assertEqual(
            set(payload["items"][0]),
            {
                "id", "product", "product_number", "product_name", "on_shelf_quantity",
                "current_store_stock", "current_store_monthly_needed", "notes", "updated_at",
                "preferred_supplier_id", "supplier_name", "supplier_number", "supplier_order_code",
                "supplier_full_name", "supplier_pack_size", "supplier_purchase_price", "supplier_names",
                "commodity_group", "commodity_group_number",
            },
        )
        self.assertNotIn("joe_quantity", payload["items"][0])
        self.assertEqual(
            self.client.patch(
                f"/api/items/{item.id}/", {"joe_quantity": 2}, content_type="application/json"
            ).status_code,
            403,
        )
        self.assertEqual(
            self.client.patch(
                f"/api/items/{item.id}/", {"on_shelf_quantity": 2}, content_type="application/json"
            ).status_code,
            200,
        )

    def test_normal_user_cannot_open_bulk_orders(self):
        self.assertEqual(self.client.get("/bulk-orders/").status_code, 403)
        self.assertEqual(self.client.get("/api/bulk-orders/").status_code, 403)

    def test_normal_user_can_delete_draft_items_but_not_finalized_items(self):
        product = Product.objects.create(
            korona_id=uuid.uuid4(), number="904", name="Delete rule", normalized_name="deleterule"
        )
        draft = OrderList.objects.create(
            store=self.store, order_date=date(2026, 7, 23), created_by=self.user
        )
        draft_item = OrderListItem.objects.create(
            order_list=draft, product=product, created_by=self.user, updated_by=self.user
        )
        self.assertEqual(self.client.delete(f"/api/items/{draft_item.id}/").status_code, 204)

        finalized = OrderList.objects.create(
            store=self.store,
            order_date=date(2026, 7, 24),
            status=OrderList.Status.FINALIZED,
            created_by=self.user,
        )
        finalized_item = OrderListItem.objects.create(
            order_list=finalized, product=product, created_by=self.user, updated_by=self.user
        )
        self.assertEqual(self.client.delete(f"/api/items/{finalized_item.id}/").status_code, 403)
        self.assertTrue(OrderListItem.objects.filter(pk=finalized_item.pk).exists())

    def test_finalized_list_is_read_only_and_shows_transfers_to_normal_users(self):
        order_list = OrderList.objects.create(
            store=self.store,
            order_date=date(2026, 7, 21),
            status=OrderList.Status.FINALIZED,
            created_by=self.user,
        )
        product = Product.objects.create(
            korona_id=uuid.uuid4(), number="902", name="Final", normalized_name="final"
        )
        item = OrderListItem.objects.create(
            order_list=order_list, product=product, created_by=self.user, updated_by=self.user
        )
        source = Store.objects.create(korona_id=uuid.uuid4(), number="2", name="Source")
        OrderItemTransfer.objects.create(item=item, from_store=source, quantity=3)

        payload = self.client.get(f"/api/orders/{order_list.id}/").json()

        self.assertFalse(payload["permissions"]["can_edit"])
        self.assertEqual(payload["items"][0]["transfers"][0]["from_store_number"], "2")
        self.assertEqual(
            self.client.patch(
                f"/api/items/{item.id}/", {"notes": "changed"}, content_type="application/json"
            ).status_code,
            403,
        )

    def test_admin_can_finalize_edit_and_delete_a_list(self):
        self.user.is_staff = True
        self.user.save(update_fields=["is_staff"])
        order_list = OrderList.objects.create(
            store=self.store, order_date=date(2026, 7, 22), created_by=self.user
        )
        product = Product.objects.create(
            korona_id=uuid.uuid4(), number="903", name="Admin Final", normalized_name="adminfinal"
        )
        item = OrderListItem.objects.create(
            order_list=order_list, product=product, created_by=self.user, updated_by=self.user
        )

        self.assertEqual(self.client.post(f"/api/orders/{order_list.id}/finalize/").status_code, 200)
        self.assertEqual(
            self.client.patch(
                f"/api/items/{item.id}/", {"joe_quantity": 4}, content_type="application/json"
            ).status_code,
            200,
        )
        self.assertEqual(self.client.delete(f"/api/orders/{order_list.id}/").status_code, 204)
        self.assertFalse(OrderList.objects.filter(pk=order_list.pk).exists())

    def test_admin_can_manage_users_but_cannot_change_superuser(self):
        self.user.is_staff = True
        self.user.save(update_fields=["is_staff"])
        superuser = get_user_model().objects.create_superuser("root-user", password="strong-pass-123")
        regular = get_user_model().objects.create_user("clerk", password="strong-pass-123")

        promoted = self.client.patch(
            f"/api/users/{regular.id}/", {"is_admin": True}, content_type="application/json"
        )

        self.assertEqual(promoted.status_code, 200)
        regular.refresh_from_db()
        self.assertTrue(regular.is_staff)
        self.assertEqual(
            self.client.patch(
                f"/api/users/{superuser.id}/", {"is_admin": False}, content_type="application/json"
            ).status_code,
            403,
        )
        self.assertEqual(self.client.delete(f"/api/users/{superuser.id}/").status_code, 403)


@override_settings(
    DELIVERY_BUCKET_ENDPOINT="https://storage.example.test",
    DELIVERY_BUCKET_NAME="proof-bucket",
    DELIVERY_BUCKET_ACCESS_KEY_ID="test-key",
    DELIVERY_BUCKET_SECRET_ACCESS_KEY="test-secret",
)
class DeliveryProofTests(TestCase):
    def setUp(self):
        self.worker = get_user_model().objects.create_user("driver", password="test-pass-123")
        self.admin = get_user_model().objects.create_user(
            "delivery-admin", password="test-pass-123", is_staff=True
        )
        self.other = get_user_model().objects.create_user("other-driver", password="test-pass-123")
        self.store = Store.objects.create(
            korona_id=uuid.uuid4(), number="101", name="Main Street", active=True
        )
        self.client.force_login(self.worker)

    def create_delivery(self, **overrides):
        values = {
            "store": self.store,
            "delivered_at": timezone.now(),
            "submitted_by": self.worker,
            "general_notes": "Received by the store manager",
        }
        values.update(overrides)
        return Delivery.objects.create(**values)

    def add_asset(self, delivery, category, name):
        return DeliveryAsset.objects.create(
            delivery=delivery,
            category=category,
            object_key=f"{delivery.storage_prefix}/{category}/{uuid.uuid4()}-{name}",
            original_filename=name,
            content_type="image/jpeg",
            size_bytes=1234,
            upload_status=DeliveryAsset.UploadStatus.UPLOADED,
            uploaded_by=self.worker,
        )

    def test_worker_can_create_draft_with_notes_and_only_sees_own_deliveries(self):
        response = self.client.post(
            "/api/deliveries/",
            {
                "store_id": self.store.id,
                "delivered_at": timezone.now().isoformat(),
                "reference_number": "INV-4402",
                "general_notes": "12 cases placed in back room",
                "issue_notes": "One wet box",
                "expected_cases": 12,
                "delivered_cases": 12,
                "damaged_cases": 1,
            },
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 201)
        delivery = Delivery.objects.get(uuid=response.json()["uuid"])
        self.assertEqual(delivery.reference_number, "INV-4402")
        self.assertEqual(delivery.issue_notes, "One wet box")
        self.assertTrue(delivery.events.filter(event_type=DeliveryEvent.EventType.CREATED).exists())

        hidden = self.create_delivery(submitted_by=self.other)
        listed = self.client.get("/api/deliveries/").json()["deliveries"]
        self.assertNotIn(str(hidden.uuid), [row["uuid"] for row in listed])

    def test_delivery_storage_prefix_is_structured_and_stable(self):
        delivery = self.create_delivery(
            delivered_at=timezone.make_aware(datetime(2026, 7, 15, 13, 30))
        )

        self.assertEqual(
            delivery.storage_prefix,
            f"deliveries/2026/07/15/store-101/delivery-{delivery.uuid}",
        )

    def test_other_worker_cannot_view_delivery_but_admin_can(self):
        delivery = self.create_delivery()
        self.client.force_login(self.other)
        self.assertEqual(self.client.get(f"/api/deliveries/{delivery.uuid}/").status_code, 403)
        self.client.force_login(self.admin)
        self.assertEqual(self.client.get(f"/api/deliveries/{delivery.uuid}/").status_code, 200)

    @patch("orders.delivery_views.presigned_upload", return_value="https://upload.example.test")
    def test_presign_validates_image_and_creates_pending_asset(self, presign):
        delivery = self.create_delivery()
        response = self.client.post(
            f"/api/deliveries/{delivery.uuid}/assets/presign/",
            {
                "category": "invoice",
                "filename": "Invoice page 1.jpg",
                "content_type": "image/jpeg",
                "size_bytes": 4000,
            },
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 201)
        asset = DeliveryAsset.objects.get(uuid=response.json()["asset_uuid"])
        self.assertEqual(asset.upload_status, DeliveryAsset.UploadStatus.PENDING)
        self.assertIn(f"{delivery.storage_prefix}/invoice/", asset.object_key)
        presign.assert_called_once()

    def test_submit_requires_both_invoice_and_box_photos(self):
        delivery = self.create_delivery()
        self.add_asset(delivery, DeliveryAsset.Category.INVOICE, "invoice.jpg")

        response = self.client.post(f"/api/deliveries/{delivery.uuid}/submit/")

        self.assertEqual(response.status_code, 400)
        self.assertIn("boxes", response.json()["detail"].lower())

    @patch("orders.delivery_views.notes_snapshot")
    def test_submission_and_admin_review_are_audited(self, snapshot):
        delivery = self.create_delivery(issue_notes="Two cases missing")
        self.add_asset(delivery, DeliveryAsset.Category.INVOICE, "invoice.jpg")
        self.add_asset(delivery, DeliveryAsset.Category.BOXES, "boxes.jpg")
        submitted = self.client.post(f"/api/deliveries/{delivery.uuid}/submit/")
        self.assertEqual(submitted.status_code, 200)
        delivery.refresh_from_db()
        self.assertEqual(delivery.status, Delivery.Status.SUBMITTED)

        self.client.force_login(self.admin)
        reviewed = self.client.post(
            f"/api/deliveries/{delivery.uuid}/review/",
            {"status": "issue_found", "admin_notes": "Contact supplier for shortage credit."},
            content_type="application/json",
        )

        self.assertEqual(reviewed.status_code, 200)
        delivery.refresh_from_db()
        self.assertEqual(delivery.status, Delivery.Status.ISSUE_FOUND)
        self.assertEqual(delivery.reviewed_by, self.admin)
        self.assertTrue(delivery.events.filter(event_type=DeliveryEvent.EventType.REVIEWED).exists())
        self.assertEqual(snapshot.call_count, 2)

    @patch("orders.delivery_views.notes_snapshot")
    def test_admin_keywords_make_delivery_searchable(self, _snapshot):
        delivery = self.create_delivery(reference_number="INV-123")
        self.client.force_login(self.admin)
        saved = self.client.put(
            f"/api/deliveries/{delivery.uuid}/keywords/",
            {"keywords": ["Holiday rush", "Supplier claim"]},
            content_type="application/json",
        )
        results = self.client.get("/api/deliveries/?q=holiday").json()["deliveries"]

        self.assertEqual(saved.status_code, 200)
        self.assertEqual([row["uuid"] for row in results], [str(delivery.uuid)])
        self.assertTrue(DeliveryKeyword.objects.filter(normalized_name="holiday rush").exists())

    def test_admin_cannot_verify_an_unsubmitted_draft(self):
        delivery = self.create_delivery()
        self.client.force_login(self.admin)
        response = self.client.post(
            f"/api/deliveries/{delivery.uuid}/review/",
            {"status": "verified", "admin_notes": "Too early"},
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 409)

    @patch("orders.delivery_views.notes_snapshot")
    def test_worker_can_answer_needs_information_with_more_notes_and_proof(self, _snapshot):
        delivery = self.create_delivery()
        self.add_asset(delivery, DeliveryAsset.Category.INVOICE, "invoice.jpg")
        self.add_asset(delivery, DeliveryAsset.Category.BOXES, "boxes.jpg")
        self.assertEqual(self.client.post(f"/api/deliveries/{delivery.uuid}/submit/").status_code, 200)
        self.client.force_login(self.admin)
        self.assertEqual(
            self.client.post(
                f"/api/deliveries/{delivery.uuid}/review/",
                {"status": "needs_info", "admin_notes": "Please identify the damaged case."},
                content_type="application/json",
            ).status_code,
            200,
        )
        self.client.force_login(self.worker)
        updated = self.client.patch(
            f"/api/deliveries/{delivery.uuid}/",
            {"issue_notes": "Damage is in the case at the left of the stack."},
            content_type="application/json",
        )
        resubmitted = self.client.post(f"/api/deliveries/{delivery.uuid}/submit/")
        self.assertEqual(updated.status_code, 200)
        self.assertEqual(resubmitted.status_code, 200)
        delivery.refresh_from_db()
        self.assertEqual(delivery.status, Delivery.Status.SUBMITTED)

    def test_only_admin_can_export_delivery_log(self):
        self.create_delivery(reference_number="INV-CSV")
        self.assertEqual(self.client.get("/api/deliveries/export.csv").status_code, 403)
        self.client.force_login(self.admin)
        exported = self.client.get("/api/deliveries/export.csv")
        self.assertEqual(exported.status_code, 200)
        self.assertIn("INV-CSV", exported.content.decode())
