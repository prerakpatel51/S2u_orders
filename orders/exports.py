from io import BytesIO
from xml.sax.saxutils import escape

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter, portrait
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


DISTRIBUTORS = {
    "joe": ("JOE", "joe_quantity"),
    "bt": ("BT", "bt_quantity"),
    "sqw": ("SQW", "sqw_quantity"),
}
EXPORT_KINDS = {"transfers", *DISTRIBUTORS}
BLUE = "2F66E8"
TRANSFER_COLORS = ["E7F2E7", "F8EEDC", "F0E4F1", "E5F1FA", "E2F1EE", "FCE8E6", "FFF4CC", "E8EAF6"]


def _items(order_list, item_ids=None):
    queryset = order_list.items.select_related("product").prefetch_related(
        "product__codes", "transfers__from_store"
    )
    if item_ids is not None:
        queryset = queryset.filter(id__in=item_ids)
    return queryset.order_by("row_order", "id")


def _barcode(item):
    code = next(iter(item.product.codes.all()), None)
    return code.code if code else ""


def _store_sort_key(store_number):
    value = str(store_number).lstrip("#")
    return (0, int(value)) if value.isdigit() else (1, value.casefold())


def bulk_order_export_response(bulk_order, file_format):
    stores = list(bulk_order.items.model._meta.apps.get_model("orders", "Store").objects.filter(active=True).order_by("number"))
    items = bulk_order.items.select_related("product").prefetch_related("product__codes", "quantities")
    headers = ["Barcode", "Product"] + [store.number for store in stores] + ["Total cases"]
    rows = []
    for item in items:
        quantities = {value.store_id: float(value.cases) for value in item.quantities.all()}
        values = [quantities.get(store.id, 0) for store in stores]
        barcode = next((code.code for code in item.product.codes.all()), "")
        rows.append([barcode, item.product.name, *values, sum(values)])
    filename = "".join(character if character.isalnum() or character in "-_" else "_" for character in bulk_order.name).strip("_") or "bulk-order"
    if file_format == "xlsx":
        workbook = Workbook(); sheet = workbook.active; sheet.title = "Bulk Order"; sheet.append(headers)
        for row in rows: sheet.append(row)
        for cell in sheet[1]: cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor=BLUE)
        sheet.freeze_panes = "C2"; sheet.auto_filter.ref = sheet.dimensions
        sheet.column_dimensions["A"].width = 15; sheet.column_dimensions["B"].width = 38
        for index in range(3, len(headers) + 1): sheet.column_dimensions[get_column_letter(index)].width = 13
        output = BytesIO(); workbook.save(output); response = HttpResponse(output.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"); response["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'; return response
    output = BytesIO(); document = SimpleDocTemplate(output, pagesize=letter, rightMargin=24, leftMargin=24, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet(); story = [Paragraph(bulk_order.name, styles["Title"]), Spacer(1, 10)]
    table = Table([headers, *rows], repeatRows=1, hAlign="LEFT"); table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F66E8")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("FONTSIZE", (0, 0), (-1, -1), 6), ("GRID", (0, 0), (-1, -1), .3, colors.HexColor("#CCD5E0")), ("VALIGN", (0, 0), (-1, -1), "MIDDLE")]))
    story.append(table); document.build(story); response = HttpResponse(output.getvalue(), content_type="application/pdf"); response["Content-Disposition"] = f'attachment; filename="{filename}.pdf"'; return response


def _transfer_color_map(rows):
    stores = dict.fromkeys(row[2] for row in rows)
    return {store: TRANSFER_COLORS[index % len(TRANSFER_COLORS)] for index, store in enumerate(stores)}


def export_spec(order_list, kind, item_ids=None):
    items = _items(order_list, item_ids)
    if kind in DISTRIBUTORS:
        label, field = DISTRIBUTORS[kind]
        rows = [
            [item.product.name, item.product.number, _barcode(item), float(getattr(item, field))]
            for item in items
            if getattr(item, field) > 0
        ]
        return f"{label} Export", ["Product Name", "Product #", "Barcode", label], rows

    if kind == "transfers":
        rows = []
        for item in items:
            for transfer in item.transfers.all():
                rows.append(
                    [item.product.number, item.product.name, f"#{transfer.from_store.number}", float(transfer.quantity)]
                )
        rows.sort(key=lambda row: (_store_sort_key(row[2]), str(row[1]).casefold(), str(row[0])))
        return "Transfer List", ["Product #", "Product Name", "From", "Bottles"], rows

    raise ValueError(f"Unsupported export kind: {kind}")


def _store_label(order_list):
    return order_list.store.name or f"Store {order_list.store.number}"


def _metadata(order_list, generated_by):
    generated = timezone.localtime().strftime("%Y-%m-%d %H:%M")
    return f"Week of {order_list.order_date:%Y-%m-%d} | Generated on {generated} by {generated_by}"


def xlsx_response(order_list, kind="order", generated_by="system", item_ids=None):
    title, headers, rows = export_spec(order_list, kind, item_ids)
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    last_column = get_column_letter(len(headers))
    ws.merge_cells(f"A1:{last_column}1")
    ws["A1"] = f"{title} - {_store_label(order_list)} (#{order_list.store.number})"
    ws["A1"].font = Font(size=18, bold=True, color=BLUE)
    ws.merge_cells(f"A2:{last_column}2")
    ws["A2"] = _metadata(order_list, generated_by)
    ws["A2"].font = Font(size=11, color="28313D")

    for column, value in enumerate(headers, 1):
        cell = ws.cell(row=4, column=column, value=value)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in rows:
        ws.append(row)

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{last_column}{max(ws.max_row, 4)}"
    if kind in DISTRIBUTORS:
        widths = [38, 15, 20, 10]
    elif kind == "transfers":
        widths = [16, 44, 16, 12]
    else:
        widths = [16, 38, 10, 10, 10, 22, 14]
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width
    transfer_colors = _transfer_color_map(rows) if kind == "transfers" else {}
    for row_number, row in enumerate(ws.iter_rows(min_row=5), 5):
        if kind == "transfers":
            fill_color = transfer_colors.get(ws.cell(row=row_number, column=3).value)
            if fill_color:
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor=fill_color)
        elif row_number % 2 == 0:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor="F3F6FB")
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = "0.###"

    buffer = BytesIO()
    wb.save(buffer)
    filename = f"{kind}-{order_list.store.number}-{order_list.order_date}.xlsx"
    response = HttpResponse(
        buffer.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def pdf_response(order_list, kind="order", generated_by="system", item_ids=None):
    title, headers, rows = export_spec(order_list, kind, item_ids)
    buffer = BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=0.45 * inch,
        leftMargin=0.45 * inch,
        topMargin=0.45 * inch,
        bottomMargin=0.45 * inch,
        title=f"{title} - {_store_label(order_list)}",
    )
    styles = getSampleStyleSheet()
    heading = ParagraphStyle(
        "ExportHeading",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=22,
        leading=27,
        textColor=colors.HexColor(f"#{BLUE}"),
        alignment=0,
        spaceAfter=8,
    )
    body = ParagraphStyle("ExportBody", parent=styles["BodyText"], fontSize=10, leading=13)
    story = [
        Paragraph(f"{title} - {_store_label(order_list)} (#{order_list.store.number})", heading),
        Paragraph(_metadata(order_list, generated_by), body),
        Spacer(1, 20),
    ]

    data = [headers]
    for row in rows:
        data.append([Paragraph(str(value), body) if isinstance(value, str) else f"{value:g}" for value in row])
    if len(data) == 1:
        data.append([Paragraph("No items", body)] + ["-"] * (len(headers) - 1))

    if kind in DISTRIBUTORS:
        widths = [3.2 * inch, 1.1 * inch, 2.0 * inch, 0.8 * inch]
    elif kind == "transfers":
        widths = [1.0 * inch, 3.75 * inch, 1.35 * inch, 1.0 * inch]
    else:
        widths = [0.8 * inch, 2.4 * inch, 0.55 * inch, 0.55 * inch, 0.55 * inch, 1.35 * inch, 0.9 * inch]
    table = Table(data, repeatRows=1, colWidths=widths)
    table_commands = [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{BLUE}")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 11),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("ALIGN", (-1, 1), (-1, -1), "RIGHT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CAD4E2")),
                ("LEFTPADDING", (0, 0), (-1, -1), 7),
                ("RIGHTPADDING", (0, 0), (-1, -1), 7),
                ("TOPPADDING", (0, 0), (-1, -1), 7),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]
    if kind == "transfers":
        transfer_colors = _transfer_color_map(rows)
        for row_number, row in enumerate(rows, 1):
            table_commands.append(
                ("BACKGROUND", (0, row_number), (-1, row_number), colors.HexColor(f"#{transfer_colors[row[2]]}"))
            )
    else:
        table_commands.append(
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F6FB")])
        )
    table.setStyle(TableStyle(table_commands))
    story.append(table)
    document.build(story)
    filename = f"{kind}-{order_list.store.number}-{order_list.order_date}.pdf"
    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _grid_export_filename(title):
    return "".join(character if character.isalnum() or character in "-_" else "_" for character in title).strip("_") or "current-grid"


def grid_pdf_response(order_list, columns, rows, orientation="landscape", generated_by="system", title=""):
    """Render the exact visible, filtered, sorted grid payload supplied by the UI."""
    title = title or _store_label(order_list)
    page_size = landscape(letter) if orientation == "landscape" else portrait(letter)
    buffer = BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=page_size,
        rightMargin=0.28 * inch,
        leftMargin=0.28 * inch,
        topMargin=0.32 * inch,
        bottomMargin=0.35 * inch,
        title=title,
    )
    available_width = page_size[0] - document.leftMargin - document.rightMargin
    raw_widths = [max(45, min(float(column.get("width") or 100), 260)) for column in columns]
    width_total = sum(raw_widths) or 1
    column_widths = [available_width * width / width_total for width in raw_widths]
    font_size = max(4.2, min(8.0, available_width / max(len(columns), 1) / 7.5))
    styles = getSampleStyleSheet()
    heading = ParagraphStyle(
        "GridExportHeading", parent=styles["Title"], fontName="Helvetica-Bold",
        fontSize=16, leading=19, textColor=colors.HexColor(f"#{BLUE}"), spaceAfter=4,
    )
    meta = ParagraphStyle("GridExportMeta", parent=styles["BodyText"], fontSize=7.5, leading=9)
    cell = ParagraphStyle(
        "GridExportCell", parent=styles["BodyText"], fontName="Helvetica",
        fontSize=font_size, leading=font_size + 1.2, wordWrap="CJK",
    )
    header = ParagraphStyle(
        "GridExportHeader", parent=cell, fontName="Helvetica-Bold",
        textColor=colors.white, alignment=1,
    )

    def paragraph(value, style):
        safe = escape(str(value or "")).replace("\n", "<br/>")
        return Paragraph(safe, style)

    headers = [paragraph(column.get("label") or column.get("id") or "", header) for column in columns]
    data = [headers]
    for row in rows:
        data.append([paragraph(row[index] if index < len(row) else "", cell) for index in range(len(columns))])
    if len(data) == 1:
        data.append([paragraph("No matching rows", cell)] + [""] * (len(columns) - 1))

    table = Table(data, repeatRows=1, colWidths=column_widths, hAlign="LEFT")
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{BLUE}")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CAD4E2")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F6FB")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story = [
        Paragraph(escape(title), heading),
        Paragraph(escape(_metadata(order_list, generated_by)), meta),
        Paragraph(f"{len(rows)} filtered row(s) | {len(columns)} visible column(s) | {orientation.title()}", meta),
        Spacer(1, 10),
        table,
    ]

    def footer(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.HexColor("#66736F"))
        canvas.drawRightString(page_size[0] - document.rightMargin, 0.17 * inch, f"Page {doc.page}")
        canvas.restoreState()

    document.build(story, onFirstPage=footer, onLaterPages=footer)
    filename = f"{_grid_export_filename(title)}-{order_list.order_date}-{orientation}.pdf"
    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="{filename}"'
    response["X-PDF-Filename"] = filename
    return response


def grid_xlsx_response(order_list, columns, rows, generated_by="system", title=""):
    title = title or _store_label(order_list)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Current Grid"
    last_column = get_column_letter(len(columns))
    sheet.merge_cells(f"A1:{last_column}1")
    sheet["A1"] = title
    sheet["A1"].font = Font(size=18, bold=True, color=BLUE)
    sheet.merge_cells(f"A2:{last_column}2")
    sheet["A2"] = _metadata(order_list, generated_by)
    sheet["A2"].font = Font(size=10, color="66736F")
    for index, column in enumerate(columns, 1):
        cell = sheet.cell(row=4, column=index, value=column.get("label") or column.get("id") or "")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        pixel_width = float(column.get("width") or 100)
        sheet.column_dimensions[get_column_letter(index)].width = max(8, min(pixel_width / 7, 42))
    for row_index, values in enumerate(rows, 5):
        for column_index in range(1, len(columns) + 1):
            value = values[column_index - 1] if column_index - 1 < len(values) else ""
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if row_index % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F3F6FB")
    sheet.freeze_panes = "A5"
    sheet.auto_filter.ref = f"A4:{last_column}{max(sheet.max_row, 4)}"
    sheet.sheet_view.showGridLines = False
    sheet.page_setup.orientation = "landscape" if len(columns) > 6 else "portrait"
    sheet.page_setup.fitToWidth = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    buffer = BytesIO()
    workbook.save(buffer)
    filename = f"{_grid_export_filename(title)}-{order_list.order_date}.xlsx"
    response = HttpResponse(
        buffer.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    response["X-XLSX-Filename"] = filename
    return response
